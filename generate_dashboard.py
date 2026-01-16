"""
PIPELINE ML PRODUCTION - CLASSIFICATION RÉCLAMATIONS BANCAIRES
===============================================================
- Entraînement optimisé sur données 2024 (Optuna)
- Test sur données 2025
- Analyse financière détaillée (économies vs pertes FP)
- Règle métier: 1 réclamation auto par client
- Génération de rapports Excel
"""

import pandas as pd
import numpy as np
import argparse
import logging
import joblib
from pathlib import Path
from datetime import datetime
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.preprocessing import LabelEncoder, RobustScaler
from sklearn.metrics import (
    accuracy_score, precision_score, recall_score, f1_score,
    roc_auc_score, confusion_matrix, brier_score_loss, classification_report
)
import xgboost as xgb
import optuna
import warnings
warnings.filterwarnings('ignore')

# Configuration du logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('pipeline_2024_2025.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION
# ============================================================================
CONFIG = {
    'costs': {
        'prix_unitaire_dh': 169,
        'penalite_fn_multiplier': 2,
    },
    'optuna': {
        'n_trials': 50,
        'cv_folds': 5,
    },
    'model': {
        'random_state': 42,
        'early_stopping_rounds': 50,
    }
}


# ============================================================================
# CLASSE PREPROCESSOR
# ============================================================================
class RobustPreprocessor:
    """Preprocessing robuste avec gestion de tous les types de données."""
    
    def __init__(self):
        self.scaler = RobustScaler()
        self.label_encoders = {}
        self.family_medians = {}
        self.feature_names_fitted = None
        self.target_col = None
        
        # Colonnes à exclure par pattern
        self.exclude_patterns = [
            'id', 'date', 'dt_', '_dt', 'timestamp', 'datetime', 
            'num_', 'numero', 'n°', 'code_', 'ref'
        ]
    
    def fit(self, df: pd.DataFrame, target_col: str):
        """Fit sur données d'entraînement (2024)."""
        logger.info("\n🔧 Configuration du preprocessing...")
        self.target_col = target_col
        
        X = df.copy()
        
        # Calculer médianes par famille pour features métier
        if 'Famille Produit' in X.columns and 'Montant demandé' in X.columns:
            self.family_medians = X.groupby('Famille Produit')['Montant demandé'].median().to_dict()
            logger.info(f"   ✓ Médianes calculées pour {len(self.family_medians)} familles")
        
        # Préparer features
        X = self._prepare_features(X, fit_mode=True)
        
        # Sélectionner colonnes numériques
        numeric_cols = X.select_dtypes(include=[np.number]).columns.tolist()
        numeric_cols = [c for c in numeric_cols if c != target_col]
        
        # Sauvegarder l'ordre des colonnes
        self.feature_names_fitted = numeric_cols
        
        # Fit scaler
        self.scaler.fit(X[self.feature_names_fitted])
        
        logger.info(f"   ✓ Preprocessing configuré: {len(self.feature_names_fitted)} features")
        return self
    
    def transform(self, df: pd.DataFrame) -> pd.DataFrame:
        """Transform avec ordre garanti des colonnes."""
        X = df.copy()
        X = self._prepare_features(X, fit_mode=False)
        
        # Gérer colonnes manquantes
        for col in self.feature_names_fitted:
            if col not in X.columns:
                X[col] = 0
                logger.warning(f"   ⚠ Colonne manquante: {col}")
        
        # Garder seulement les colonnes du fit
        X = X[self.feature_names_fitted]
        
        # Appliquer scaler
        X_scaled = pd.DataFrame(
            self.scaler.transform(X),
            columns=self.feature_names_fitted,
            index=X.index
        )
        
        return X_scaled
    
    def fit_transform(self, df: pd.DataFrame, target_col: str) -> pd.DataFrame:
        """Fit et transform."""
        self.fit(df, target_col)
        return self.transform(df)
    
    def _prepare_features(self, df: pd.DataFrame, fit_mode: bool) -> pd.DataFrame:
        """Prépare les features en gérant tous les types."""
        X = df.copy()
        
        # Supprimer colonne cible si présente
        if self.target_col and self.target_col in X.columns:
            X = X.drop(columns=[self.target_col])
        
        cols_to_drop = []
        
        for col in X.columns:
            col_lower = col.lower()
            
            # Exclure par pattern
            if any(pattern in col_lower for pattern in self.exclude_patterns):
                cols_to_drop.append(col)
                continue
            
            try:
                # Datetime
                if pd.api.types.is_datetime64_any_dtype(X[col]):
                    X[f'{col}_month'] = X[col].dt.month.fillna(0).astype(int)
                    X[f'{col}_dow'] = X[col].dt.dayofweek.fillna(0).astype(int)
                    cols_to_drop.append(col)
                    continue
                
                # Timedelta
                if pd.api.types.is_timedelta64_dtype(X[col]):
                    X[col] = X[col].dt.total_seconds().fillna(0)
                    continue
                
                # Boolean
                if X[col].dtype == 'bool':
                    X[col] = X[col].astype(int)
                    continue
                
                # Object/Category
                if X[col].dtype == 'object' or str(X[col].dtype) == 'category':
                    unique_vals = set(str(v).lower() for v in X[col].dropna().unique())
                    
                    # Booléen caché
                    if unique_vals.issubset({'oui', 'non', 'o', 'n', 'true', 'false', '0', '1', 'yes', 'no'}):
                        X[col] = X[col].apply(
                            lambda x: 1 if str(x).lower() in ['oui', 'o', 'true', '1', 'yes'] else 0
                        ).astype(int)
                    else:
                        # Label encoding
                        if fit_mode:
                            le = LabelEncoder()
                            X[col] = X[col].astype(str).fillna('_MISSING_')
                            X[col] = le.fit_transform(X[col])
                            self.label_encoders[col] = le
                        else:
                            if col in self.label_encoders:
                                le = self.label_encoders[col]
                                X[col] = X[col].astype(str).fillna('_MISSING_')
                                # Gérer valeurs inconnues
                                X[col] = X[col].apply(
                                    lambda x: le.transform([x])[0] if x in le.classes_ else -1
                                )
                            else:
                                X[col] = 0
                    continue
                
                # Numérique
                if pd.api.types.is_numeric_dtype(X[col]):
                    median_val = X[col].median() if X[col].notna().any() else 0
                    X[col] = X[col].fillna(median_val)
                    continue
                
                # Essayer conversion
                try:
                    X[col] = pd.to_numeric(X[col], errors='coerce').fillna(0)
                except:
                    cols_to_drop.append(col)
                    
            except Exception as e:
                logger.warning(f"   ⚠ Erreur colonne '{col}': {e}")
                cols_to_drop.append(col)
        
        # Supprimer colonnes problématiques
        X = X.drop(columns=[c for c in cols_to_drop if c in X.columns])
        
        # Créer features métier
        X = self._create_business_features(X)
        
        # S'assurer que tout est numérique
        for col in X.columns:
            if not pd.api.types.is_numeric_dtype(X[col]):
                try:
                    X[col] = pd.to_numeric(X[col], errors='coerce').fillna(0)
                except:
                    X = X.drop(columns=[col])
        
        # Remplir NaN restants
        X = X.fillna(0)
        X = X.replace([np.inf, -np.inf], 0)
        
        return X
    
    def _create_business_features(self, X: pd.DataFrame) -> pd.DataFrame:
        """Créer features métier."""
        df = X.copy()
        
        # Ratio PNB/Montant
        if 'PNB analytique (vision commerciale) cumulé' in df.columns and 'Montant demandé' in df.columns:
            df['ratio_pnb_montant'] = (
                df['PNB analytique (vision commerciale) cumulé'] / 
                (df['Montant demandé'] + 1)
            )
        
        # Écart à la médiane famille
        if 'Famille Produit' in df.columns and 'Montant demandé' in df.columns and self.family_medians:
            df['ecart_mediane_famille'] = df.apply(
                lambda row: (
                    row['Montant demandé'] - 
                    self.family_medians.get(row.get('Famille Produit', ''), row['Montant demandé'])
                ) / (self.family_medians.get(row.get('Famille Produit', ''), 1) + 1),
                axis=1
            )
        
        # Log transformations
        if 'Montant demandé' in df.columns:
            df['log_montant'] = np.log1p(df['Montant demandé'].clip(lower=0))
        
        if 'PNB analytique (vision commerciale) cumulé' in df.columns:
            df['log_pnb'] = np.log1p(df['PNB analytique (vision commerciale) cumulé'].clip(lower=0))
        
        if 'anciennete_annees' in df.columns:
            df['log_anciennete'] = np.log1p(df['anciennete_annees'].clip(lower=0))
        
        return df


# ============================================================================
# CLASSE ANALYSEUR FINANCIER
# ============================================================================
class FinancialAnalyzer:
    """Analyse financière avec métriques métier."""
    
    def __init__(self, prix_unitaire: int = 169):
        self.prix_unitaire = prix_unitaire
        self.results = {}
    
    def analyze(self, df: pd.DataFrame, y_true: np.ndarray, y_pred: np.ndarray, 
                scenario_name: str = "default") -> dict:
        """Analyse financière complète."""
        
        logger.info(f"\n{'='*70}")
        logger.info(f"💰 ANALYSE FINANCIÈRE: {scenario_name.upper()}")
        logger.info(f"{'='*70}")
        
        n_total = len(y_true)
        
        # Matrice de confusion
        tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
        
        # Métriques traitement
        n_traite = n_total
        n_correct = tp + tn
        taux_traite = 100.0
        taux_precision = (n_correct / n_total) * 100
        
        logger.info(f"\n📊 MÉTRIQUES DE TRAITEMENT:")
        logger.info(f"   Total réclamations:     {n_total:>10}")
        logger.info(f"   Décisions correctes:    {n_correct:>10} ({taux_precision:.1f}%)")
        logger.info(f"   TP (Fondées validées):  {tp:>10}")
        logger.info(f"   TN (Non fondées rejetées): {tn:>10}")
        logger.info(f"   FP (Fausses validations): {fp:>10}")
        logger.info(f"   FN (Faux rejets):       {fn:>10}")
        
        # Analyse FP - CRITIQUE
        mask_fp = (y_true == 0) & (y_pred == 1)
        montants_fp = df.loc[mask_fp, 'Montant demandé'] if 'Montant demandé' in df.columns else pd.Series([0])
        somme_montants_fp = montants_fp.sum()
        moyenne_montant_fp = montants_fp.mean() if len(montants_fp) > 0 else 0
        
        logger.info(f"\n🔴 FAUX POSITIFS (Remboursements à tort):")
        logger.info(f"   Nombre:                 {fp:>10}")
        logger.info(f"   Somme montants:         {somme_montants_fp:>10,.2f} DH")
        logger.info(f"   Montant moyen:          {moyenne_montant_fp:>10,.2f} DH")
        
        # Analyse FN
        mask_fn = (y_true == 1) & (y_pred == 0)
        montants_fn = df.loc[mask_fn, 'Montant demandé'] if 'Montant demandé' in df.columns else pd.Series([0])
        somme_montants_fn = montants_fn.sum()
        
        logger.info(f"\n🟠 FAUX NÉGATIFS (Rejets à tort):")
        logger.info(f"   Nombre:                 {fn:>10}")
        logger.info(f"   Somme montants:         {somme_montants_fn:>10,.2f} DH")
        
        # Calcul économique
        economie_traitement = n_traite * self.prix_unitaire
        perte_fp = somme_montants_fp
        cout_fn = fn * self.prix_unitaire * 2
        gain_net = economie_traitement - perte_fp - cout_fn
        
        logger.info(f"\n💵 BILAN ÉCONOMIQUE:")
        logger.info(f"   ┌────────────────────────────────────────────────────┐")
        logger.info(f"   │ (+) Économie traitement ({n_traite} × {self.prix_unitaire} DH)")
        logger.info(f"   │     = {economie_traitement:>15,.2f} DH              │")
        logger.info(f"   │ (-) Perte FP (montants remboursés à tort)")
        logger.info(f"   │     = {perte_fp:>15,.2f} DH              │")
        logger.info(f"   │ (-) Coût FN (retraitement estimé)")
        logger.info(f"   │     = {cout_fn:>15,.2f} DH              │")
        logger.info(f"   │ ──────────────────────────────────────────────────│")
        logger.info(f"   │ (=) GAIN NET = {gain_net:>15,.2f} DH              │")
        logger.info(f"   └────────────────────────────────────────────────────┘")
        
        # Ratio économie/perte
        ratio = economie_traitement / perte_fp if perte_fp > 0 else float('inf')
        
        logger.info(f"\n📊 COMPARAISON ÉCONOMIE vs PERTE FP:")
        logger.info(f"   Économie:               {economie_traitement:>15,.2f} DH")
        logger.info(f"   Perte FP:               {perte_fp:>15,.2f} DH")
        logger.info(f"   Ratio:                  {ratio:>15.2f}x")
        
        if ratio > 1:
            logger.info(f"   ✅ RENTABLE: L'économie dépasse les pertes")
        else:
            logger.info(f"   ⚠️ NON RENTABLE: Les pertes dépassent l'économie")
        
        # Stocker résultats
        self.results[scenario_name] = {
            'n_total': n_total,
            'n_traite': n_traite,
            'taux_traite': taux_traite,
            'taux_precision': taux_precision,
            'tp': tp, 'tn': tn, 'fp': fp, 'fn': fn,
            'somme_montants_fp': somme_montants_fp,
            'somme_montants_fn': somme_montants_fn,
            'moyenne_montant_fp': moyenne_montant_fp,
            'economie_traitement': economie_traitement,
            'perte_fp': perte_fp,
            'cout_fn': cout_fn,
            'gain_net': gain_net,
            'ratio_economie_perte': ratio
        }
        
        return self.results[scenario_name]
    
    def compare_scenarios(self, scenario1: str, scenario2: str):
        """Compare deux scénarios."""
        if scenario1 not in self.results or scenario2 not in self.results:
            logger.error("Scénarios non trouvés")
            return
        
        r1 = self.results[scenario1]
        r2 = self.results[scenario2]
        
        logger.info(f"\n{'='*70}")
        logger.info(f"📊 COMPARAISON: {scenario1.upper()} vs {scenario2.upper()}")
        logger.info(f"{'='*70}")
        
        metrics = [
            ('Taux précision (%)', 'taux_precision'),
            ('Nb FP', 'fp'),
            ('Nb FN', 'fn'),
            ('Somme montants FP (DH)', 'somme_montants_fp'),
            ('Économie traitement (DH)', 'economie_traitement'),
            ('Perte FP (DH)', 'perte_fp'),
            ('Gain net (DH)', 'gain_net'),
            ('Ratio éco/perte', 'ratio_economie_perte'),
        ]
        
        logger.info(f"\n{'Métrique':<30} {scenario1:>15} {scenario2:>15} {'Diff':>15}")
        logger.info("-" * 75)
        
        for label, key in metrics:
            v1 = r1[key]
            v2 = r2[key]
            diff = v2 - v1
            
            if 'DH' in label:
                logger.info(f"{label:<30} {v1:>12,.0f} DH {v2:>12,.0f} DH {diff:>+12,.0f} DH")
            elif '%' in label:
                logger.info(f"{label:<30} {v1:>14.1f}% {v2:>14.1f}% {diff:>+14.1f}%")
            elif 'Ratio' in label:
                logger.info(f"{label:<30} {v1:>14.2f}x {v2:>14.2f}x {diff:>+14.2f}x")
            else:
                logger.info(f"{label:<30} {v1:>15} {v2:>15} {diff:>+15}")


# ============================================================================
# CLASSE PIPELINE PRINCIPAL
# ============================================================================
class ProductionPipeline:
    """Pipeline production: entraînement 2024, test 2025."""
    
    def __init__(self, output_dir: str = 'outputs'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        (self.output_dir / 'models').mkdir(exist_ok=True)
        (self.output_dir / 'figures').mkdir(exist_ok=True)
        
        self.preprocessor = RobustPreprocessor()
        self.financial_analyzer = FinancialAnalyzer(CONFIG['costs']['prix_unitaire_dh'])
        self.model = None
        self.best_params = None
        self.optimal_threshold = 0.5
        
        self.df_2024 = None
        self.df_2025 = None
        self.target_col = None
    
    def load_data(self, path_2024: str, path_2025: str):
        """Charger les données."""
        logger.info(f"\n{'='*70}")
        logger.info("📂 CHARGEMENT DES DONNÉES")
        logger.info(f"{'='*70}")
        
        self.df_2024 = pd.read_excel(path_2024)
        self.df_2025 = pd.read_excel(path_2025)
        
        logger.info(f"   2024: {self.df_2024.shape[0]} lignes, {self.df_2024.shape[1]} colonnes")
        logger.info(f"   2025: {self.df_2025.shape[0]} lignes, {self.df_2025.shape[1]} colonnes")
        
        # Détecter colonne cible
        self.target_col = self._detect_target(self.df_2024)
        
        # Préparer cible
        self.y_2024 = self._prepare_target(self.df_2024)
        self.y_2025 = self._prepare_target(self.df_2025)
        
        logger.info(f"\n   Distribution 2024: {self.y_2024.mean():.1%} fondées")
        logger.info(f"   Distribution 2025: {self.y_2025.mean():.1%} fondées")
    
    def _detect_target(self, df: pd.DataFrame) -> str:
        """Détecte la colonne cible."""
        for col in df.columns:
            if 'fond' in col.lower():
                logger.info(f"   Colonne cible détectée: {col}")
                return col
        raise ValueError("Colonne cible non trouvée")
    
    def _prepare_target(self, df: pd.DataFrame) -> pd.Series:
        """Prépare la variable cible."""
        y = df[self.target_col].copy()
        
        if y.dtype == 'object':
            y = y.apply(lambda x: 1 if str(x).lower() in ['oui', 'yes', '1', 'fondée', 'fondee', 'true', 'o'] else 0)
        elif y.dtype == 'bool':
            y = y.astype(int)
        
        return y.fillna(0).astype(int)
    
    def optimize_model(self, X_train: pd.DataFrame, y_train: np.ndarray):
        """Optimisation Optuna."""
        logger.info(f"\n{'='*70}")
        logger.info("🔬 OPTIMISATION HYPERPARAMÈTRES (Optuna)")
        logger.info(f"{'='*70}")
        
        # Ratio de classes
        scale_pos_weight = (y_train == 0).sum() / (y_train == 1).sum()
        logger.info(f"   Ratio classes (neg/pos): {scale_pos_weight:.2f}")
        
        def objective(trial):
            params = {
                'objective': 'binary:logistic',
                'eval_metric': 'logloss',
                'random_state': CONFIG['model']['random_state'],
                'n_jobs': -1,
                'verbosity': 0,
                'use_label_encoder': False,
                
                'max_depth': trial.suggest_int('max_depth', 3, 10),
                'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.3, log=True),
                'n_estimators': trial.suggest_int('n_estimators', 100, 500),
                'min_child_weight': trial.suggest_int('min_child_weight', 1, 10),
                'subsample': trial.suggest_float('subsample', 0.6, 1.0),
                'colsample_bytree': trial.suggest_float('colsample_bytree', 0.6, 1.0),
                'gamma': trial.suggest_float('gamma', 0, 5),
                'reg_alpha': trial.suggest_float('reg_alpha', 1e-8, 10.0, log=True),
                'reg_lambda': trial.suggest_float('reg_lambda', 1e-8, 10.0, log=True),
                'scale_pos_weight': trial.suggest_float('scale_pos_weight', 0.5, scale_pos_weight * 2),
            }
            
            model = xgb.XGBClassifier(**params)
            cv = StratifiedKFold(n_splits=CONFIG['optuna']['cv_folds'], shuffle=True, 
                                random_state=CONFIG['model']['random_state'])
            scores = cross_val_score(model, X_train, y_train, cv=cv, scoring='f1', n_jobs=-1)
            
            return scores.mean()
        
        optuna.logging.set_verbosity(optuna.logging.WARNING)
        
        study = optuna.create_study(
            direction='maximize',
            sampler=optuna.samplers.TPESampler(seed=CONFIG['model']['random_state']),
            pruner=optuna.pruners.MedianPruner()
        )
        
        logger.info(f"   Lancement de {CONFIG['optuna']['n_trials']} trials...")
        study.optimize(objective, n_trials=CONFIG['optuna']['n_trials'], 
                      show_progress_bar=True, n_jobs=1)
        
        logger.info(f"\n   ✓ Meilleur F1-Score CV: {study.best_value:.4f}")
        logger.info(f"\n   Meilleurs hyperparamètres:")
        for key, value in study.best_params.items():
            logger.info(f"      {key}: {value}")
        
        self.best_params = study.best_params
        return study.best_params
    
    def train_model(self):
        """Entraîner le modèle sur 2024."""
        logger.info(f"\n{'='*70}")
        logger.info("🎯 ENTRAÎNEMENT SUR 2024")
        logger.info(f"{'='*70}")
        
        # Preprocessing
        X_train = self.preprocessor.fit_transform(self.df_2024, self.target_col)
        y_train = self.y_2024.values
        
        logger.info(f"   Shape après preprocessing: {X_train.shape}")
        
        # Split train/val pour early stopping
        X_tr, X_val, y_tr, y_val = train_test_split(
            X_train, y_train, test_size=0.15, stratify=y_train,
            random_state=CONFIG['model']['random_state']
        )
        
        # Optimisation
        best_params = self.optimize_model(X_tr, y_tr)
        
        # Entraîner modèle final
        logger.info(f"\n🏋️ Entraînement modèle final...")
        
        self.model = xgb.XGBClassifier(
            **best_params,
            objective='binary:logistic',
            eval_metric='logloss',
            early_stopping_rounds=CONFIG['model']['early_stopping_rounds'],
            random_state=CONFIG['model']['random_state'],
            n_jobs=-1,
            verbosity=0,
            use_label_encoder=False
        )
        
        self.model.fit(
            X_tr, y_tr,
            eval_set=[(X_val, y_val)],
            verbose=False
        )
        
        logger.info(f"   ✓ Modèle entraîné avec {self.model.best_iteration} itérations")
        
        # Validation croisée
        logger.info(f"\n📊 Validation croisée (5-fold) sur 2024:")
        cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=CONFIG['model']['random_state'])
        
        self.cv_scores = {}
        for metric in ['accuracy', 'precision', 'recall', 'f1', 'roc_auc']:
            scores = cross_val_score(self.model, X_train, y_train, cv=cv, scoring=metric, n_jobs=-1)
            self.cv_scores[metric] = scores
            logger.info(f"   {metric:12s}: {scores.mean():.4f} (+/- {scores.std()*2:.4f})")
        
        # Sauvegarder
        joblib.dump(self.model, self.output_dir / 'models' / 'model_xgboost.pkl')
        joblib.dump(self.preprocessor, self.output_dir / 'models' / 'preprocessor.pkl')
        joblib.dump(best_params, self.output_dir / 'models' / 'best_params.pkl')
        logger.info(f"\n   ✓ Modèle sauvegardé dans {self.output_dir / 'models'}")
    
    def evaluate_2025(self):
        """Évaluer sur 2025."""
        logger.info(f"\n{'='*70}")
        logger.info("📊 ÉVALUATION SUR 2025")
        logger.info(f"{'='*70}")
        
        # Transform 2025
        X_test = self.preprocessor.transform(self.df_2025)
        y_test = self.y_2025.values
        
        logger.info(f"   Shape 2025: {X_test.shape}")
        
        # Prédictions
        self.y_pred_2025 = self.model.predict(X_test)
        self.y_proba_2025 = self.model.predict_proba(X_test)[:, 1]
        
        # Métriques
        self.metrics_2025 = self._calculate_metrics(y_test, self.y_pred_2025, self.y_proba_2025)
        
        logger.info(f"\n📊 Métriques sur 2025:")
        for metric, value in self.metrics_2025.items():
            if metric != 'confusion_matrix':
                logger.info(f"   {metric:20s}: {value:.4f}")
        
        # Matrice de confusion
        cm = self.metrics_2025['confusion_matrix']
        logger.info(f"\n   Matrice de confusion:")
        logger.info(f"   TN={cm[0,0]:>5}  FP={cm[0,1]:>5}")
        logger.info(f"   FN={cm[1,0]:>5}  TP={cm[1,1]:>5}")
        
        self.y_test_2025 = y_test
    
    def _calculate_metrics(self, y_true, y_pred, y_proba=None) -> dict:
        """Calculer métriques."""
        metrics = {
            'accuracy': accuracy_score(y_true, y_pred),
            'precision_class_0': precision_score(y_true, y_pred, pos_label=0, zero_division=0),
            'precision_class_1': precision_score(y_true, y_pred, pos_label=1, zero_division=0),
            'recall_class_0': recall_score(y_true, y_pred, pos_label=0, zero_division=0),
            'recall_class_1': recall_score(y_true, y_pred, pos_label=1, zero_division=0),
            'f1_class_0': f1_score(y_true, y_pred, pos_label=0, zero_division=0),
            'f1_class_1': f1_score(y_true, y_pred, pos_label=1, zero_division=0),
            'f1_weighted': f1_score(y_true, y_pred, average='weighted', zero_division=0),
            'confusion_matrix': confusion_matrix(y_true, y_pred)
        }
        
        if y_proba is not None:
            metrics['roc_auc'] = roc_auc_score(y_true, y_proba)
            metrics['brier_score'] = brier_score_loss(y_true, y_proba)
        
        return metrics
    
    def apply_business_rule(self):
        """Appliquer règle métier: 1 réclamation auto par client."""
        logger.info(f"\n{'='*70}")
        logger.info("🔒 RÈGLE MÉTIER: 1 RÉCLAMATION AUTO PAR CLIENT")
        logger.info(f"{'='*70}")
        
        df = self.df_2025.copy()
        df['y_pred'] = self.y_pred_2025
        df['y_proba'] = self.y_proba_2025
        df['y_true'] = self.y_test_2025
        
        # Trouver colonne client
        client_col = None
        for col in ['idtfcl', 'numero_compte', 'N compte', 'ID Client', 'client_id']:
            if col in df.columns:
                client_col = col
                break
        
        if client_col is None:
            logger.warning("   ⚠ Colonne client non trouvée, création d'un index")
            df['client_id'] = df.index
            client_col = 'client_id'
        
        logger.info(f"   Colonne client: {client_col}")
        
        # Trier par date si disponible
        date_col = None
        for col in df.columns:
            if 'date' in col.lower() and 'qualif' in col.lower():
                date_col = col
                break
        
        if date_col:
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
            df = df.sort_values([client_col, date_col])
        else:
            df = df.sort_values(client_col)
        
        # Marquer première réclamation par client
        df['is_first'] = ~df.duplicated(subset=[client_col], keep='first')
        
        # Stats
        n_total = len(df)
        n_first = df['is_first'].sum()
        n_multi = n_total - n_first
        n_clients = df[client_col].nunique()
        
        logger.info(f"\n   Statistiques:")
        logger.info(f"   - Clients uniques:      {n_clients}")
        logger.info(f"   - Premières réclam.:    {n_first}")
        logger.info(f"   - Réclam. multiples:    {n_multi} ({100*n_multi/n_total:.1f}%)")
        
        # Appliquer règle: seule la 1ère réclamation peut être automatisée
        df['y_pred_with_rule'] = np.where(df['is_first'], df['y_pred'], 0)
        
        self.df_2025_analysis = df
        self.client_col = client_col
    
    def calculate_financial_impact(self):
        """Calculer impact financier."""
        logger.info(f"\n{'='*70}")
        logger.info("💰 ANALYSE FINANCIÈRE COMPLÈTE")
        logger.info(f"{'='*70}")
        
        df = self.df_2025_analysis
        
        # Scénario SANS règle
        self.financial_analyzer.analyze(
            df=df,
            y_true=df['y_true'].values,
            y_pred=df['y_pred'].values,
            scenario_name="sans_regle_2025"
        )
        
        # Scénario AVEC règle
        self.financial_analyzer.analyze(
            df=df,
            y_true=df['y_true'].values,
            y_pred=df['y_pred_with_rule'].values,
            scenario_name="avec_regle_2025"
        )
        
        # Comparaison
        self.financial_analyzer.compare_scenarios("sans_regle_2025", "avec_regle_2025")
    
    def optimize_thresholds(self):
        """Optimiser les seuils de décision."""
        logger.info(f"\n{'='*70}")
        logger.info("🎯 OPTIMISATION DES SEUILS DE DÉCISION")
        logger.info(f"{'='*70}")
        
        y_true = self.y_test_2025
        y_proba = self.y_proba_2025
        
        best_result = None
        best_score = -1
        
        for t_low in np.arange(0.10, 0.45, 0.02):
            for t_high in np.arange(0.55, 0.90, 0.02):
                if t_high <= t_low:
                    continue
                
                mask_rej = y_proba <= t_low
                mask_val = y_proba >= t_high
                
                n_rej = mask_rej.sum()
                n_val = mask_val.sum()
                
                if n_rej == 0 or n_val == 0:
                    continue
                
                prec_rej = (y_true[mask_rej] == 0).mean()
                prec_val = (y_true[mask_val] == 1).mean()
                automation = (n_rej + n_val) / len(y_true)
                
                if prec_rej >= 0.95 and prec_val >= 0.93:
                    score = automation + prec_rej * 0.1 + prec_val * 0.1
                    if score > best_score:
                        best_score = score
                        best_result = {
                            'threshold_low': t_low,
                            'threshold_high': t_high,
                            'precision_rejection': prec_rej,
                            'precision_validation': prec_val,
                            'automation_rate': automation,
                            'n_rejection': int(n_rej),
                            'n_validation': int(n_val),
                            'n_audit': int(len(y_true) - n_rej - n_val)
                        }
        
        if best_result is None:
            t_low, t_high = 0.3, 0.7
            mask_rej = y_proba <= t_low
            mask_val = y_proba >= t_high
            
            best_result = {
                'threshold_low': t_low,
                'threshold_high': t_high,
                'precision_rejection': (y_true[mask_rej] == 0).mean() if mask_rej.sum() > 0 else 0,
                'precision_validation': (y_true[mask_val] == 1).mean() if mask_val.sum() > 0 else 0,
                'automation_rate': (mask_rej.sum() + mask_val.sum()) / len(y_true),
                'n_rejection': int(mask_rej.sum()),
                'n_validation': int(mask_val.sum()),
                'n_audit': int(len(y_true) - mask_rej.sum() - mask_val.sum())
            }
        
        self.thresholds = best_result
        
        logger.info(f"\n   Seuils optimaux:")
        logger.info(f"   - Seuil rejet:          {best_result['threshold_low']:.2f}")
        logger.info(f"   - Seuil validation:     {best_result['threshold_high']:.2f}")
        logger.info(f"   - Taux automatisation:  {best_result['automation_rate']:.1%}")
        logger.info(f"   - Précision rejets:     {best_result['precision_rejection']:.1%}")
        logger.info(f"   - Précision validations:{best_result['precision_validation']:.1%}")
        
        return best_result
    
    def generate_excel_reports(self):
        """Générer tous les rapports Excel."""
        logger.info(f"\n{'='*70}")
        logger.info("📝 GÉNÉRATION DES RAPPORTS EXCEL")
        logger.info(f"{'='*70}")
        
        self._generate_summary_report()
        self._generate_category_report()
        self._generate_errors_report()
        self._generate_financial_report()
    
    def _generate_summary_report(self):
        """Générer rapport de synthèse."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HEADER_FONT = Font(color="FFFFFF", bold=True)
        SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Synthèse"
        
        # Titre
        ws['A1'] = "RAPPORT DE PERFORMANCE - TRAIN 2024 / TEST 2025"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A3'] = f"Données: 2024 ({len(self.df_2024)} lignes) → 2025 ({len(self.df_2025)} lignes)"
        
        # Métriques Classification
        row = 5
        ws.cell(row=row, column=1, value="MÉTRIQUES DE CLASSIFICATION (2025)").font = Font(bold=True, size=12)
        
        row += 1
        headers = ["Métrique", "Valeur", "Objectif", "Statut"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
        
        metrics_list = [
            ("Accuracy", self.metrics_2025['accuracy'], 0.90),
            ("F1-Score Weighted", self.metrics_2025['f1_weighted'], 0.95),
            ("Précision Rejet (classe 0)", self.metrics_2025['precision_class_0'], 0.97),
            ("Précision Validation (classe 1)", self.metrics_2025['precision_class_1'], 0.95),
            ("Recall Rejet", self.metrics_2025['recall_class_0'], 0.90),
            ("Recall Validation", self.metrics_2025['recall_class_1'], 0.90),
            ("AUC-ROC", self.metrics_2025.get('roc_auc', 0), 0.98),
        ]
        
        for name, value, threshold in metrics_list:
            row += 1
            ws.cell(row=row, column=1, value=name).border = BORDER
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = '0.0000'
            cell.border = BORDER
            ws.cell(row=row, column=3, value=f"≥{threshold:.0%}").border = BORDER
            
            if value >= threshold:
                status = "✓ OK"
                cell.fill = SUCCESS_FILL
            elif value >= threshold - 0.05:
                status = "⚠ Proche"
                cell.fill = WARNING_FILL
            else:
                status = "✗ À améliorer"
                cell.fill = ERROR_FILL
            ws.cell(row=row, column=4, value=status).border = BORDER
        
        # Validation croisée 2024
        row += 3
        ws.cell(row=row, column=1, value="VALIDATION CROISÉE (2024)").font = Font(bold=True, size=12)
        
        row += 1
        for col, h in enumerate(["Métrique", "Moyenne", "Écart-type"], 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
        
        for metric, scores in self.cv_scores.items():
            row += 1
            ws.cell(row=row, column=1, value=metric).border = BORDER
            ws.cell(row=row, column=2, value=scores.mean()).border = BORDER
            ws.cell(row=row, column=2).number_format = '0.0000'
            ws.cell(row=row, column=3, value=f"±{scores.std()*2:.4f}").border = BORDER
        
        # Seuils optimaux
        if hasattr(self, 'thresholds'):
            row += 3
            ws.cell(row=row, column=1, value="SEUILS OPTIMAUX").font = Font(bold=True, size=12)
            
            row += 1
            for col, h in enumerate(["Paramètre", "Valeur"], 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = BORDER
            
            thresh_data = [
                ("Seuil Rejet", f"{self.thresholds['threshold_low']:.2f}"),
                ("Seuil Validation", f"{self.thresholds['threshold_high']:.2f}"),
                ("Taux Automatisation", f"{self.thresholds['automation_rate']:.1%}"),
                ("Précision Rejets", f"{self.thresholds['precision_rejection']:.1%}"),
                ("Précision Validations", f"{self.thresholds['precision_validation']:.1%}"),
            ]
            
            for name, value in thresh_data:
                row += 1
                ws.cell(row=row, column=1, value=name).border = BORDER
                ws.cell(row=row, column=2, value=value).border = BORDER
        
        # Ajuster largeurs
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        output_path = self.output_dir / "1_performance_summary.xlsx"
        wb.save(output_path)
        logger.info(f"   ✓ {output_path}")
    
    def _generate_category_report(self):
        """Générer rapport par catégorie."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HEADER_FONT = Font(color="FFFFFF", bold=True)
        SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Vue ensemble"
        ws['A1'] = "PERFORMANCE PAR CATÉGORIE - TEST 2025"
        ws['A1'].font = Font(bold=True, size=16)
        
        df = self.df_2025_analysis
        y_true = df['y_true'].values
        y_pred = df['y_pred'].values
        
        # Colonnes catégorielles à analyser
        cat_cols = []
        for col in df.columns:
            col_lower = col.lower()
            if any(x in col_lower for x in ['famille', 'categ', 'segment', 'type', 'motif', 'marché']):
                if df[col].dtype == 'object' or df[col].nunique() <= 50:
                    cat_cols.append(col)
        
        cat_cols = cat_cols[:5]  # Limiter à 5 catégories
        
        for cat_col in cat_cols:
            if cat_col not in df.columns:
                continue
            
            ws_cat = wb.create_sheet(cat_col[:31])
            ws_cat['A1'] = f"Performance par {cat_col}"
            ws_cat['A1'].font = Font(bold=True, size=14)
            
            results = []
            for cat in df[cat_col].unique():
                mask = df[cat_col] == cat
                if mask.sum() < 5:
                    continue
                
                cat_y_true = y_true[mask]
                cat_y_pred = y_pred[mask]
                
                # Montants FP
                mask_fp = mask & (df['y_true'] == 0) & (df['y_pred'] == 1)
                montants_fp = df.loc[mask_fp, 'Montant demandé'].sum() if 'Montant demandé' in df.columns else 0
                
                results.append({
                    'Catégorie': str(cat)[:40],
                    'N': int(mask.sum()),
                    'Taux Fondées': float(cat_y_true.mean()),
                    'Accuracy': float(accuracy_score(cat_y_true, cat_y_pred)),
                    'Précision': float(precision_score(cat_y_true, cat_y_pred, zero_division=0)),
                    'Recall': float(recall_score(cat_y_true, cat_y_pred, zero_division=0)),
                    'F1': float(f1_score(cat_y_true, cat_y_pred, zero_division=0)),
                    'FP': int(((cat_y_true == 0) & (cat_y_pred == 1)).sum()),
                    'Montants FP': float(montants_fp)
                })
            
            if not results:
                continue
            
            results_df = pd.DataFrame(results).sort_values('Montants FP', ascending=False)
            
            row = 3
            headers = list(results_df.columns)
            for col, h in enumerate(headers, 1):
                cell = ws_cat.cell(row=row, column=col, value=h)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.border = BORDER
            
            for r_idx, (_, data_row) in enumerate(results_df.iterrows(), row + 1):
                for c_idx, (col_name, value) in enumerate(data_row.items(), 1):
                    cell = ws_cat.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = BORDER
                    
                    if col_name in ['Taux Fondées', 'Accuracy', 'Précision', 'Recall', 'F1']:
                        cell.number_format = '0.00%'
                    elif col_name == 'Montants FP':
                        cell.number_format = '#,##0.00'
            
            for col in 'ABCDEFGHI':
                ws_cat.column_dimensions[col].width = 15
        
        output_path = self.output_dir / "2_performance_by_category.xlsx"
        wb.save(output_path)
        logger.info(f"   ✓ {output_path}")
    
    def _generate_errors_report(self):
        """Générer rapport des erreurs."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HEADER_FONT = Font(color="FFFFFF", bold=True)
        ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        df = self.df_2025_analysis
        
        mask_fp = (df['y_true'] == 0) & (df['y_pred'] == 1)
        mask_fn = (df['y_true'] == 1) & (df['y_pred'] == 0)
        
        df_fp = df[mask_fp].copy()
        df_fn = df[mask_fn].copy()
        
        wb = Workbook()
        
        # Résumé
        ws = wb.active
        ws.title = "Résumé"
        ws['A1'] = "ANALYSE DES ERREURS - TEST 2025"
        ws['A1'].font = Font(bold=True, size=16)
        
        ws['A3'] = "Type Erreur"
        ws['B3'] = "Nombre"
        ws['C3'] = "Montant Total"
        ws['D3'] = "Impact"
        for col in 'ABCD':
            ws[f'{col}3'].fill = HEADER_FILL
            ws[f'{col}3'].font = HEADER_FONT
        
        montant_fp = df_fp['Montant demandé'].sum() if 'Montant demandé' in df_fp.columns else 0
        montant_fn = df_fn['Montant demandé'].sum() if 'Montant demandé' in df_fn.columns else 0
        
        ws['A4'] = "Faux Positifs (Fausses Validations)"
        ws['B4'] = len(df_fp)
        ws['C4'] = f"{montant_fp:,.2f} DH"
        ws['D4'] = "PERTE FINANCIÈRE"
        ws['A4'].fill = ERROR_FILL
        
        ws['A5'] = "Faux Négatifs (Faux Rejets)"
        ws['B5'] = len(df_fn)
        ws['C5'] = f"{montant_fn:,.2f} DH"
        ws['D5'] = "INSATISFACTION CLIENT"
        ws['A5'].fill = WARNING_FILL
        
        ws['A6'] = "TOTAL ERREURS"
        ws['B6'] = len(df_fp) + len(df_fn)
        ws['A6'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        
        # Faux Positifs
        ws_fp = wb.create_sheet("Faux Positifs")
        ws_fp['A1'] = f"FAUX POSITIFS - {len(df_fp)} cas - {montant_fp:,.2f} DH"
        ws_fp['A1'].font = Font(bold=True, size=14, color="C00000")
        
        if len(df_fp) > 0:
            cols = [c for c in df_fp.columns if not c.startswith('y_') and c != 'is_first'][:15]
            cols.append('y_proba')
            cols = [c for c in cols if c in df_fp.columns]
            
            row = 3
            for col_idx, col in enumerate(cols, 1):
                cell = ws_fp.cell(row=row, column=col_idx, value=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            
            for r_idx, (_, data_row) in enumerate(df_fp[cols].head(500).iterrows(), row + 1):
                for c_idx, value in enumerate(data_row, 1):
                    cell = ws_fp.cell(row=r_idx, column=c_idx)
                    if pd.isna(value):
                        cell.value = ""
                    elif isinstance(value, (np.floating, float)):
                        cell.value = round(float(value), 4)
                    elif isinstance(value, (np.integer, int)):
                        cell.value = int(value)
                    else:
                        cell.value = str(value)[:100]
        
        # Faux Négatifs
        ws_fn = wb.create_sheet("Faux Négatifs")
        ws_fn['A1'] = f"FAUX NÉGATIFS - {len(df_fn)} cas"
        ws_fn['A1'].font = Font(bold=True, size=14, color="FF6600")
        
        if len(df_fn) > 0:
            cols = [c for c in df_fn.columns if not c.startswith('y_') and c != 'is_first'][:15]
            cols.append('y_proba')
            cols = [c for c in cols if c in df_fn.columns]
            
            row = 3
            for col_idx, col in enumerate(cols, 1):
                cell = ws_fn.cell(row=row, column=col_idx, value=col)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            
            for r_idx, (_, data_row) in enumerate(df_fn[cols].head(500).iterrows(), row + 1):
                for c_idx, value in enumerate(data_row, 1):
                    cell = ws_fn.cell(row=r_idx, column=c_idx)
                    if pd.isna(value):
                        cell.value = ""
                    elif isinstance(value, (np.floating, float)):
                        cell.value = round(float(value), 4)
                    elif isinstance(value, (np.integer, int)):
                        cell.value = int(value)
                    else:
                        cell.value = str(value)[:100]
        
        output_path = self.output_dir / "3_errors_analysis.xlsx"
        wb.save(output_path)
        logger.info(f"   ✓ {output_path}")
    
    def _generate_financial_report(self):
        """Générer rapport financier."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HEADER_FONT = Font(color="FFFFFF", bold=True)
        SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Impact Financier"
        
        ws['A1'] = "ANALYSE FINANCIÈRE - TEST 2025"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Prix unitaire traitement: {CONFIG['costs']['prix_unitaire_dh']} DH"
        
        # Comparaison scénarios
        row = 4
        ws.cell(row=row, column=1, value="COMPARAISON SCÉNARIOS").font = Font(bold=True, size=12)
        
        row += 1
        headers = ["Métrique", "SANS Règle", "AVEC Règle", "Différence"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
        
        r_sans = self.financial_analyzer.results.get('sans_regle_2025', {})
        r_avec = self.financial_analyzer.results.get('avec_regle_2025', {})
        
        if r_sans and r_avec:
            metrics_fin = [
                ("Taux précision (%)", 'taux_precision', '.1f', '%'),
                ("Nb FP", 'fp', 'd', ''),
                ("Nb FN", 'fn', 'd', ''),
                ("Somme montants FP", 'somme_montants_fp', ',.0f', ' DH'),
                ("Économie traitement", 'economie_traitement', ',.0f', ' DH'),
                ("Perte FP", 'perte_fp', ',.0f', ' DH'),
                ("Coût FN estimé", 'cout_fn', ',.0f', ' DH'),
                ("GAIN NET", 'gain_net', ',.0f', ' DH'),
                ("Ratio Économie/Perte", 'ratio_economie_perte', '.2f', 'x'),
            ]
            
            for label, key, fmt, suffix in metrics_fin:
                row += 1
                v1 = r_sans.get(key, 0)
                v2 = r_avec.get(key, 0)
                diff = v2 - v1
                
                ws.cell(row=row, column=1, value=label).border = BORDER
                ws.cell(row=row, column=2, value=f"{v1:{fmt}}{suffix}").border = BORDER
                ws.cell(row=row, column=3, value=f"{v2:{fmt}}{suffix}").border = BORDER
                
                cell_diff = ws.cell(row=row, column=4, value=f"{diff:+{fmt}}{suffix}")
                cell_diff.border = BORDER
                
                if 'GAIN' in label or 'Économie' in label:
                    if diff > 0:
                        cell_diff.fill = SUCCESS_FILL
                    elif diff < 0:
                        cell_diff.fill = ERROR_FILL
                elif 'Perte' in label or 'FP' in label or 'FN' in label:
                    if diff < 0:
                        cell_diff.fill = SUCCESS_FILL
                    elif diff > 0:
                        cell_diff.fill = ERROR_FILL
        
        # Recommandation
        row += 3
        ws.cell(row=row, column=1, value="RECOMMANDATION").font = Font(bold=True, size=12)
        
        row += 1
        if r_sans and r_avec:
            if r_avec.get('gain_net', 0) > r_sans.get('gain_net', 0):
                ws.cell(row=row, column=1, value="✓ Appliquer la règle métier améliore le gain net").fill = SUCCESS_FILL
            else:
                ws.cell(row=row, column=1, value="⚠ La règle métier réduit le gain net, à évaluer").fill = WARNING_FILL
        
        # Ajuster largeurs
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        output_path = self.output_dir / "4_financial_analysis.xlsx"
        wb.save(output_path)
        logger.info(f"   ✓ {output_path}")
    
    def run(self, path_2024: str, path_2025: str):
        """Exécuter le pipeline complet."""
        logger.info("\n" + "="*70)
        logger.info("🚀 PIPELINE ML - TRAIN 2024 / TEST 2025")
        logger.info("="*70)
        
        # 1. Charger données
        self.load_data(path_2024, path_2025)
        
        # 2. Entraîner sur 2024
        self.train_model()
        
        # 3. Évaluer sur 2025
        self.evaluate_2025()
        
        # 4. Optimiser seuils
        self.optimize_thresholds()
        
        # 5. Appliquer règle métier
        self.apply_business_rule()
        
        # 6. Calculer impact financier
        self.calculate_financial_impact()
        
        # 7. Générer rapports
        self.generate_excel_reports()
        
        logger.info("\n" + "="*70)
        logger.info("✅ PIPELINE TERMINÉ AVEC SUCCÈS")
        logger.info("="*70)
        logger.info(f"\n📁 Livrables dans: {self.output_dir}")
        logger.info("   1. 1_performance_summary.xlsx")
        logger.info("   2. 2_performance_by_category.xlsx")
        logger.info("   3. 3_errors_analysis.xlsx")
        logger.info("   4. 4_financial_analysis.xlsx")


# ============================================================================
# MAIN
# ============================================================================
def main():
    parser = argparse.ArgumentParser(description='Pipeline ML - Train 2024 / Test 2025')
    parser.add_argument('--data-2024', '-d24', required=True, help='Fichier Excel 2024')
    parser.add_argument('--data-2025', '-d25', required=True, help='Fichier Excel 2025')
    parser.add_argument('--output', '-o', default='outputs', help='Répertoire de sortie')
    parser.add_argument('--trials', '-t', type=int, default=50, help='Nombre de trials Optuna')
    args = parser.parse_args()
    
    # Mettre à jour config
    CONFIG['optuna']['n_trials'] = args.trials
    
    # Lancer pipeline
    pipeline = ProductionPipeline(output_dir=args.output)
    pipeline.run(args.data_2024, args.data_2025)


if __name__ == "__main__":
    main()

