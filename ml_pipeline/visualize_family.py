"""
VISUALISATIONS PAR FAMILLE - Script dédié
Génère des visualisations détaillées par famille de produit
Usage: python visualize_family.py
"""
import sys
sys.path.append('src')

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
from sklearn.metrics import confusion_matrix, accuracy_score
import warnings
warnings.filterwarnings('ignore')

# Configuration
sns.set_style('whitegrid')
plt.rcParams['figure.figsize'] = (20, 12)


def load_data_and_predictions():
    """Charger les données et les prédictions"""
    print("\n" + "="*80)
    print("📂 CHARGEMENT DES DONNÉES")
    print("="*80)

    # Charger les données 2025
    df_2025 = pd.read_excel('data/raw/reclamations_2025.xlsx')
    print(f"✅ Données 2025: {len(df_2025)} réclamations")

    # Pour l'instant, on va simuler des prédictions
    # Dans la vraie utilisation, vous chargerez les prédictions d'un modèle
    # Exemple: y_pred = joblib.load('outputs/production/models/predictions_2025.pkl')

    # Simulation temporaire (à remplacer par vos vraies prédictions)
    np.random.seed(42)
    y_true = df_2025['Fondee'].values
    # Simuler des prédictions avec ~99% accuracy
    y_pred = y_true.copy()
    errors = np.random.choice(len(y_true), size=int(len(y_true) * 0.01), replace=False)
    y_pred[errors] = 1 - y_pred[errors]

    print(f"✅ Prédictions chargées")
    print(f"   Accuracy globale: {accuracy_score(y_true, y_pred):.2%}")

    return df_2025, y_true, y_pred


def analyze_by_family(df_2025, y_true, y_pred):
    """Analyser les performances par famille"""
    print("\n" + "="*80)
    print("📊 ANALYSE PAR FAMILLE")
    print("="*80)

    families = df_2025['Famille Produit'].unique()
    results = []

    for family in families:
        mask = df_2025['Famille Produit'] == family

        if mask.sum() == 0:
            continue

        y_t = y_true[mask]
        y_p = y_pred[mask]

        # Métriques
        n_total = len(y_t)
        n_fondees = (y_t == 1).sum()
        n_non_fondees = (y_t == 0).sum()

        accuracy = accuracy_score(y_t, y_p)

        tp = ((y_t == 1) & (y_p == 1)).sum()
        fp = ((y_t == 0) & (y_p == 1)).sum()
        tn = ((y_t == 0) & (y_p == 0)).sum()
        fn = ((y_t == 1) & (y_p == 0)).sum()

        results.append({
            'Famille': str(family)[:40],
            'N_Total': n_total,
            'N_Fondées': n_fondees,
            'N_Non_Fondées': n_non_fondees,
            'Pct_Fondées': 100 * n_fondees / n_total,
            'Accuracy': accuracy,
            'TP': tp,
            'FP': fp,
            'TN': tn,
            'FN': fn,
            'Erreurs_Total': fp + fn
        })

    df_results = pd.DataFrame(results)
    df_results = df_results.sort_values('N_Total', ascending=False)

    print(f"✅ Analyse complétée pour {len(families)} familles")
    return df_results


def create_visualizations(df_2025, y_true, y_pred, df_results):
    """Créer toutes les visualisations"""
    print("\n" + "="*80)
    print("📊 GÉNÉRATION DES VISUALISATIONS")
    print("="*80)

    # Créer la figure principale avec plusieurs subplots
    fig = plt.figure(figsize=(20, 14))
    gs = fig.add_gridspec(4, 3, hspace=0.35, wspace=0.3)

    # ========== 1. MATRICE DE CONFUSION GLOBALE ==========
    ax1 = fig.add_subplot(gs[0, 0])
    cm = confusion_matrix(y_true, y_pred)

    # Heatmap avec annotations
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', ax=ax1,
                xticklabels=['Non Fondée', 'Fondée'],
                yticklabels=['Non Fondée', 'Fondée'],
                cbar_kws={'label': 'Nombre de cas'},
                annot_kws={'size': 14, 'weight': 'bold'})

    ax1.set_xlabel('Prédiction', fontweight='bold', fontsize=12)
    ax1.set_ylabel('Réalité', fontweight='bold', fontsize=12)
    ax1.set_title('Matrice de Confusion Globale', fontweight='bold', fontsize=13)

    # Ajouter les pourcentages
    tn, fp, fn, tp = cm.ravel()
    total = tn + fp + fn + tp
    ax1.text(0.5, -0.15, f'TN: {100*tn/total:.1f}%  |  FP: {100*fp/total:.1f}%  |  FN: {100*fn/total:.1f}%  |  TP: {100*tp/total:.1f}%',
             ha='center', transform=ax1.transAxes, fontsize=10, style='italic')

    # ========== 2. ACCURACY PAR FAMILLE (TOP 15) ==========
    ax2 = fig.add_subplot(gs[0, 1:])
    top_15 = df_results.head(15)

    colors_acc = ['#27ae60' if acc >= 0.98 else '#f39c12' if acc >= 0.95 else '#e74c3c'
                  for acc in top_15['Accuracy']]

    bars = ax2.barh(range(len(top_15)), top_15['Accuracy'] * 100, color=colors_acc, alpha=0.8)
    ax2.set_yticks(range(len(top_15)))
    ax2.set_yticklabels(top_15['Famille'], fontsize=9)
    ax2.set_xlabel('Accuracy (%)', fontweight='bold', fontsize=11)
    ax2.set_title('Accuracy par Famille (Top 15 par volume)', fontweight='bold', fontsize=13)
    ax2.grid(True, alpha=0.3, axis='x')
    ax2.set_xlim([90, 100])

    # Ajouter les valeurs et le volume
    for i, (idx, row) in enumerate(top_15.iterrows()):
        acc_val = row['Accuracy'] * 100
        volume = row['N_Total']
        ax2.text(acc_val + 0.3, i, f'{acc_val:.1f}% (n={volume})',
                va='center', fontsize=8, fontweight='bold')

    # ========== 3. DISTRIBUTION DES ERREURS PAR FAMILLE ==========
    ax3 = fig.add_subplot(gs[1, :])
    top_10 = df_results.head(10)

    x = np.arange(len(top_10))
    width = 0.35

    bars1 = ax3.bar(x - width/2, top_10['FP'], width, label='Faux Positifs (FP)',
                    color='#e74c3c', alpha=0.8, edgecolor='black', linewidth=0.5)
    bars2 = ax3.bar(x + width/2, top_10['FN'], width, label='Faux Négatifs (FN)',
                    color='#e67e22', alpha=0.8, edgecolor='black', linewidth=0.5)

    ax3.set_xticks(x)
    ax3.set_xticklabels(top_10['Famille'], rotation=45, ha='right', fontsize=9)
    ax3.set_ylabel('Nombre d\'erreurs', fontweight='bold', fontsize=11)
    ax3.set_title('Distribution des Erreurs par Famille (Top 10)', fontweight='bold', fontsize=13)
    ax3.legend(loc='upper right', fontsize=10)
    ax3.grid(True, alpha=0.3, axis='y')

    # Ajouter les valeurs sur les barres
    for bars in [bars1, bars2]:
        for bar in bars:
            height = bar.get_height()
            if height > 0:
                ax3.text(bar.get_x() + bar.get_width()/2., height + 0.3,
                        f'{int(height)}', ha='center', va='bottom', fontsize=8, fontweight='bold')

    # ========== 4. VOLUME PAR FAMILLE ==========
    ax4 = fig.add_subplot(gs[2, 0])
    top_10_vol = df_results.head(10)

    bars = ax4.barh(range(len(top_10_vol)), top_10_vol['N_Total'],
                    color='#3498db', alpha=0.8, edgecolor='black', linewidth=0.5)
    ax4.set_yticks(range(len(top_10_vol)))
    ax4.set_yticklabels(top_10_vol['Famille'], fontsize=9)
    ax4.set_xlabel('Nombre de réclamations', fontweight='bold', fontsize=11)
    ax4.set_title('Volume par Famille', fontweight='bold', fontsize=13)
    ax4.grid(True, alpha=0.3, axis='x')

    # Ajouter les valeurs
    for i, val in enumerate(top_10_vol['N_Total']):
        ax4.text(val + 50, i, f'{int(val)}', va='center', fontsize=9, fontweight='bold')

    # ========== 5. TAUX DE FONDÉES PAR FAMILLE ==========
    ax5 = fig.add_subplot(gs[2, 1])
    top_10_fond = df_results.head(10)

    colors_fond = ['#9b59b6' if pct >= 50 else '#3498db' for pct in top_10_fond['Pct_Fondées']]

    bars = ax5.barh(range(len(top_10_fond)), top_10_fond['Pct_Fondées'],
                    color=colors_fond, alpha=0.8, edgecolor='black', linewidth=0.5)
    ax5.set_yticks(range(len(top_10_fond)))
    ax5.set_yticklabels(top_10_fond['Famille'], fontsize=9)
    ax5.set_xlabel('% Fondées', fontweight='bold', fontsize=11)
    ax5.set_title('Taux de Réclamations Fondées', fontweight='bold', fontsize=13)
    ax5.grid(True, alpha=0.3, axis='x')
    ax5.axvline(x=50, color='red', linestyle='--', alpha=0.5, linewidth=2)

    # Ajouter les valeurs
    for i, val in enumerate(top_10_fond['Pct_Fondées']):
        ax5.text(val + 1, i, f'{val:.1f}%', va='center', fontsize=9, fontweight='bold')

    # ========== 6. TAUX D'ERREUR PAR FAMILLE ==========
    ax6 = fig.add_subplot(gs[2, 2])
    top_10_err = df_results.head(10)

    taux_erreur = 100 * top_10_err['Erreurs_Total'] / top_10_err['N_Total']
    colors_err = ['#27ae60' if err <= 2 else '#f39c12' if err <= 5 else '#e74c3c'
                  for err in taux_erreur]

    bars = ax6.barh(range(len(top_10_err)), taux_erreur,
                    color=colors_err, alpha=0.8, edgecolor='black', linewidth=0.5)
    ax6.set_yticks(range(len(top_10_err)))
    ax6.set_yticklabels(top_10_err['Famille'], fontsize=9)
    ax6.set_xlabel('Taux d\'erreur (%)', fontweight='bold', fontsize=11)
    ax6.set_title('Taux d\'Erreur par Famille', fontweight='bold', fontsize=13)
    ax6.grid(True, alpha=0.3, axis='x')

    # Ajouter les valeurs
    for i, val in enumerate(taux_erreur):
        ax6.text(val + 0.1, i, f'{val:.1f}%', va='center', fontsize=9, fontweight='bold')

    # ========== 7. COMPARAISON TP/TN/FP/FN PAR FAMILLE ==========
    ax7 = fig.add_subplot(gs[3, :])
    top_8 = df_results.head(8)

    x = np.arange(len(top_8))
    width = 0.2

    bars1 = ax7.bar(x - 1.5*width, top_8['TP'], width, label='TP (Vrais Positifs)',
                    color='#27ae60', alpha=0.8, edgecolor='black', linewidth=0.5)
    bars2 = ax7.bar(x - 0.5*width, top_8['TN'], width, label='TN (Vrais Négatifs)',
                    color='#3498db', alpha=0.8, edgecolor='black', linewidth=0.5)
    bars3 = ax7.bar(x + 0.5*width, top_8['FP'], width, label='FP (Faux Positifs)',
                    color='#e74c3c', alpha=0.8, edgecolor='black', linewidth=0.5)
    bars4 = ax7.bar(x + 1.5*width, top_8['FN'], width, label='FN (Faux Négatifs)',
                    color='#e67e22', alpha=0.8, edgecolor='black', linewidth=0.5)

    ax7.set_xticks(x)
    ax7.set_xticklabels(top_8['Famille'], rotation=45, ha='right', fontsize=9)
    ax7.set_ylabel('Nombre de cas', fontweight='bold', fontsize=11)
    ax7.set_title('Décomposition Complète par Famille (TP/TN/FP/FN)', fontweight='bold', fontsize=13)
    ax7.legend(loc='upper right', fontsize=10, ncol=4)
    ax7.grid(True, alpha=0.3, axis='y')

    # Titre principal
    plt.suptitle('ANALYSE DÉTAILLÉE PAR FAMILLE DE PRODUIT - 2025',
                fontsize=18, fontweight='bold', y=0.995)

    # Sauvegarder
    output_dir = Path('outputs/production/figures')
    output_dir.mkdir(parents=True, exist_ok=True)

    output_path = output_dir / 'visualizations_family.png'
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    print(f"✅ Sauvegardé: visualizations_family.png")

    plt.close()

    # Créer une deuxième figure avec les familles à problème
    create_problem_families_viz(df_results)


def create_problem_families_viz(df_results):
    """Créer une visualisation focalisée sur les familles à problème"""
    print("\n📊 Génération des visualisations des familles à problème...")

    # Identifier les familles avec le plus d'erreurs
    df_problems = df_results.nlargest(10, 'Erreurs_Total')

    fig, axes = plt.subplots(2, 2, figsize=(16, 10))
    fig.suptitle('FAMILLES NÉCESSITANT ATTENTION (Plus d\'erreurs)',
                fontsize=16, fontweight='bold')

    # 1. Accuracy des familles à problème
    ax = axes[0, 0]
    colors = ['#e74c3c' if acc < 0.95 else '#f39c12' if acc < 0.98 else '#27ae60'
              for acc in df_problems['Accuracy']]
    bars = ax.barh(range(len(df_problems)), df_problems['Accuracy'] * 100, color=colors, alpha=0.8)
    ax.set_yticks(range(len(df_problems)))
    ax.set_yticklabels(df_problems['Famille'], fontsize=9)
    ax.set_xlabel('Accuracy (%)', fontweight='bold')
    ax.set_title('Accuracy des Familles à Problème', fontweight='bold')
    ax.grid(True, alpha=0.3, axis='x')

    for i, val in enumerate(df_problems['Accuracy'] * 100):
        ax.text(val + 0.2, i, f'{val:.1f}%', va='center', fontsize=8, fontweight='bold')

    # 2. Nombre total d'erreurs
    ax = axes[0, 1]
    ax.barh(range(len(df_problems)), df_problems['Erreurs_Total'],
            color='#e74c3c', alpha=0.8, edgecolor='black', linewidth=0.5)
    ax.set_yticks(range(len(df_problems)))
    ax.set_yticklabels(df_problems['Famille'], fontsize=9)
    ax.set_xlabel('Nombre d\'erreurs (FP + FN)', fontweight='bold')
    ax.set_title('Total des Erreurs', fontweight='bold')
    ax.grid(True, alpha=0.3, axis='x')

    for i, val in enumerate(df_problems['Erreurs_Total']):
        ax.text(val + 1, i, f'{int(val)}', va='center', fontsize=8, fontweight='bold')

    # 3. Distribution FP vs FN
    ax = axes[1, 0]
    x = np.arange(len(df_problems))
    width = 0.35

    ax.bar(x - width/2, df_problems['FP'], width, label='FP',
           color='#e74c3c', alpha=0.8, edgecolor='black', linewidth=0.5)
    ax.bar(x + width/2, df_problems['FN'], width, label='FN',
           color='#e67e22', alpha=0.8, edgecolor='black', linewidth=0.5)

    ax.set_xticks(x)
    ax.set_xticklabels(df_problems['Famille'], rotation=45, ha='right', fontsize=8)
    ax.set_ylabel('Nombre', fontweight='bold')
    ax.set_title('FP vs FN', fontweight='bold')
    ax.legend()
    ax.grid(True, alpha=0.3, axis='y')

    # 4. Taux d'erreur
    ax = axes[1, 1]
    taux_err = 100 * df_problems['Erreurs_Total'] / df_problems['N_Total']
    colors = ['#e74c3c' if t > 5 else '#f39c12' if t > 2 else '#27ae60' for t in taux_err]

    ax.barh(range(len(df_problems)), taux_err, color=colors, alpha=0.8, edgecolor='black', linewidth=0.5)
    ax.set_yticks(range(len(df_problems)))
    ax.set_yticklabels(df_problems['Famille'], fontsize=9)
    ax.set_xlabel('Taux d\'erreur (%)', fontweight='bold')
    ax.set_title('Taux d\'Erreur', fontweight='bold')
    ax.grid(True, alpha=0.3, axis='x')
    ax.axvline(x=2, color='green', linestyle='--', alpha=0.5, label='Objectif: 2%')
    ax.axvline(x=5, color='orange', linestyle='--', alpha=0.5, label='Seuil: 5%')
    ax.legend(fontsize=8)

    for i, val in enumerate(taux_err):
        ax.text(val + 0.2, i, f'{val:.1f}%', va='center', fontsize=8, fontweight='bold')

    plt.tight_layout()

    output_dir = Path('outputs/production/figures')
    output_path = output_dir / 'problem_families.png'
    plt.savefig(output_path, dpi=300, bbox_inches='tight')
    print(f"✅ Sauvegardé: problem_families.png")

    plt.close()


def save_summary_table(df_results):
    """Sauvegarder un tableau récapitulatif en Excel"""
    print("\n📝 Génération du tableau récapitulatif Excel...")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Résumé par Famille"

    # Titre
    ws['A1'] = "RÉSUMÉ PAR FAMILLE DE PRODUIT - 2025"
    ws['A1'].font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells('A1:K1')

    # Headers
    row = 3
    headers = ['Famille', 'Volume', 'N Fondées', 'N Non Fondées', '% Fondées',
               'Accuracy', 'TP', 'FP', 'TN', 'FN', 'Erreurs Total']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')

    # Données
    for r_idx, (_, data_row) in enumerate(df_results.iterrows(), row + 1):
        ws.cell(row=r_idx, column=1, value=data_row['Famille']).border = BORDER
        ws.cell(row=r_idx, column=2, value=int(data_row['N_Total'])).border = BORDER
        ws.cell(row=r_idx, column=3, value=int(data_row['N_Fondées'])).border = BORDER
        ws.cell(row=r_idx, column=4, value=int(data_row['N_Non_Fondées'])).border = BORDER

        cell = ws.cell(row=r_idx, column=5, value=data_row['Pct_Fondées'] / 100)
        cell.number_format = '0.0%'
        cell.border = BORDER

        # Accuracy avec code couleur
        cell = ws.cell(row=r_idx, column=6, value=data_row['Accuracy'])
        cell.number_format = '0.00%'
        cell.border = BORDER
        if data_row['Accuracy'] >= 0.98:
            cell.fill = SUCCESS_FILL
        elif data_row['Accuracy'] >= 0.95:
            cell.fill = WARNING_FILL
        else:
            cell.fill = ERROR_FILL

        ws.cell(row=r_idx, column=7, value=int(data_row['TP'])).border = BORDER
        ws.cell(row=r_idx, column=8, value=int(data_row['FP'])).border = BORDER
        ws.cell(row=r_idx, column=9, value=int(data_row['TN'])).border = BORDER
        ws.cell(row=r_idx, column=10, value=int(data_row['FN'])).border = BORDER
        ws.cell(row=r_idx, column=11, value=int(data_row['Erreurs_Total'])).border = BORDER

    # Ajuster largeurs
    ws.column_dimensions['A'].width = 40
    for col in 'BCDEFGHIJK':
        ws.column_dimensions[col].width = 14

    output_dir = Path('outputs/production')
    output_path = output_dir / 'resume_famille.xlsx'
    wb.save(output_path)
    print(f"✅ Sauvegardé: resume_famille.xlsx")


def main():
    """Fonction principale"""
    print("="*80)
    print("GÉNÉRATION DES VISUALISATIONS PAR FAMILLE")
    print("="*80)

    # Charger les données
    df_2025, y_true, y_pred = load_data_and_predictions()

    # Analyser par famille
    df_results = analyze_by_family(df_2025, y_true, y_pred)

    # Afficher le résumé
    print("\n" + "="*80)
    print("📋 RÉSUMÉ PAR FAMILLE (Top 10)")
    print("="*80)
    print(df_results.head(10)[['Famille', 'N_Total', 'Accuracy', 'Erreurs_Total']].to_string(index=False))

    # Créer les visualisations
    create_visualizations(df_2025, y_true, y_pred, df_results)

    # Sauvegarder le tableau Excel
    save_summary_table(df_results)

    print("\n" + "="*80)
    print("✅ GÉNÉRATION TERMINÉE")
    print("="*80)
    print("\n📂 Fichiers générés:")
    print("   - outputs/production/figures/visualizations_family.png")
    print("   - outputs/production/figures/problem_families.png")
    print("   - outputs/production/resume_famille.xlsx")


if __name__ == '__main__':
    main()
