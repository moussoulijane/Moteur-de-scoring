"""
ANALYSE PAR FAMILLE - Accuracy et Matrice de Confusion
Génère des résumés détaillés par famille de produit et matrice de confusion globale
"""
import sys
sys.path.append('src')

import pandas as pd
import numpy as np
import joblib
from pathlib import Path
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import confusion_matrix, classification_report
import warnings
warnings.filterwarnings('ignore')

# Configuration
sns.set_style('whitegrid')
plt.rcParams['figure.figsize'] = (16, 10)


class FamilyAnalyzer:
    """Analyse les performances par famille de produit"""

    def __init__(self):
        self.results = {}

    def load_results(self):
        """Charger les résultats du meilleur modèle"""
        print("\n" + "="*80)
        print("📂 CHARGEMENT DES RÉSULTATS")
        print("="*80)

        # Charger les données 2025
        self.df_2025 = pd.read_excel('data/raw/reclamations_2025.xlsx')

        # Charger le meilleur modèle (supposons que c'est sauvegardé)
        # Si vous n'avez pas encore exécuté model_comparison, vous devrez d'abord le faire

        print(f"✅ Données 2025 chargées: {len(self.df_2025)} réclamations")

        # Vérifier si on a les colonnes nécessaires
        if 'Famille Produit' not in self.df_2025.columns:
            print("⚠️  Colonne 'Famille Produit' non trouvée")
            return False

        if 'Fondee' not in self.df_2025.columns:
            print("⚠️  Colonne 'Fondee' non trouvée")
            return False

        return True

    def load_predictions(self, y_true, y_pred, y_prob=None):
        """Charger les prédictions manuellement"""
        self.y_true = y_true
        self.y_pred = y_pred
        self.y_prob = y_prob

        print(f"✅ Prédictions chargées: {len(y_true)} cas")

    def analyze_by_family(self):
        """Analyser les performances par famille"""
        print("\n" + "="*80)
        print("📊 ANALYSE PAR FAMILLE DE PRODUIT")
        print("="*80)

        families = self.df_2025['Famille Produit'].unique()
        results_by_family = []

        for family in families:
            # Masque pour cette famille
            mask = self.df_2025['Famille Produit'] == family

            if mask.sum() == 0:
                continue

            # Données pour cette famille
            y_true_fam = self.y_true[mask]
            y_pred_fam = self.y_pred[mask]

            # Métriques
            n_total = len(y_true_fam)
            n_fondees = (y_true_fam == 1).sum()
            n_non_fondees = (y_true_fam == 0).sum()

            # Accuracy globale
            accuracy = (y_true_fam == y_pred_fam).mean()

            # True Positives, False Positives, True Negatives, False Negatives
            tp = ((y_true_fam == 1) & (y_pred_fam == 1)).sum()
            fp = ((y_true_fam == 0) & (y_pred_fam == 1)).sum()
            tn = ((y_true_fam == 0) & (y_pred_fam == 0)).sum()
            fn = ((y_true_fam == 1) & (y_pred_fam == 0)).sum()

            # Accuracy par classe
            if n_fondees > 0:
                acc_fondees = tp / n_fondees
            else:
                acc_fondees = 0

            if n_non_fondees > 0:
                acc_non_fondees = tn / n_non_fondees
            else:
                acc_non_fondees = 0

            # Precision et Recall
            if (tp + fp) > 0:
                precision = tp / (tp + fp)
            else:
                precision = 0

            if (tp + fn) > 0:
                recall = tp / (tp + fn)
            else:
                recall = 0

            # F1-Score
            if (precision + recall) > 0:
                f1 = 2 * precision * recall / (precision + recall)
            else:
                f1 = 0

            results_by_family.append({
                'Famille': str(family)[:40],
                'N_Total': n_total,
                'N_Fondées': n_fondees,
                'N_Non_Fondées': n_non_fondees,
                'Pct_Fondées': 100 * n_fondees / n_total if n_total > 0 else 0,
                'Accuracy_Globale': accuracy,
                'Accuracy_Fondées': acc_fondees,
                'Accuracy_Non_Fondées': acc_non_fondees,
                'Precision': precision,
                'Recall': recall,
                'F1_Score': f1,
                'TP': tp,
                'FP': fp,
                'TN': tn,
                'FN': fn
            })

        self.results_by_family = pd.DataFrame(results_by_family)
        self.results_by_family = self.results_by_family.sort_values('N_Total', ascending=False)

        print(f"\n✅ Analyse complétée pour {len(families)} familles")
        return self.results_by_family

    def display_summary(self):
        """Afficher le résumé"""
        print("\n" + "="*80)
        print("📋 RÉSUMÉ PAR FAMILLE")
        print("="*80)

        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', 40)

        print("\n" + self.results_by_family.to_string(index=False))

        # Statistiques globales
        print("\n" + "="*80)
        print("📊 STATISTIQUES GLOBALES")
        print("="*80)
        print(f"\nAccuracy moyenne: {self.results_by_family['Accuracy_Globale'].mean():.2%}")
        print(f"Precision moyenne: {self.results_by_family['Precision'].mean():.2%}")
        print(f"Recall moyen: {self.results_by_family['Recall'].mean():.2%}")
        print(f"F1-Score moyen: {self.results_by_family['F1_Score'].mean():.2%}")

        print(f"\nMeilleure famille (Accuracy): {self.results_by_family.iloc[self.results_by_family['Accuracy_Globale'].argmax()]['Famille']} "
              f"({self.results_by_family['Accuracy_Globale'].max():.2%})")
        print(f"Pire famille (Accuracy): {self.results_by_family.iloc[self.results_by_family['Accuracy_Globale'].argmin()]['Famille']} "
              f"({self.results_by_family['Accuracy_Globale'].min():.2%})")

    def generate_confusion_matrix_global(self):
        """Générer la matrice de confusion globale"""
        print("\n" + "="*80)
        print("🔢 MATRICE DE CONFUSION GLOBALE")
        print("="*80)

        cm = confusion_matrix(self.y_true, self.y_pred)

        print("\nMatrice de confusion:")
        print(f"                  Prédit: Non Fondée    Prédit: Fondée")
        print(f"Réel: Non Fondée        {cm[0,0]:6d}              {cm[0,1]:6d}")
        print(f"Réel: Fondée            {cm[1,0]:6d}              {cm[1,1]:6d}")

        # Calculs détaillés
        tn, fp, fn, tp = cm.ravel()

        print(f"\nDétails:")
        print(f"  True Negatives (TN):  {tn:6d} - Correct: Non Fondée → Non Fondée")
        print(f"  False Positives (FP): {fp:6d} - Erreur: Non Fondée → Fondée (COÛT FINANCIER)")
        print(f"  False Negatives (FN): {fn:6d} - Erreur: Fondée → Non Fondée (CLIENT INSATISFAIT)")
        print(f"  True Positives (TP):  {tp:6d} - Correct: Fondée → Fondée")

        total = tn + fp + fn + tp
        print(f"\nTaux d'erreur:")
        print(f"  Taux FP: {100*fp/total:.2f}%")
        print(f"  Taux FN: {100*fn/total:.2f}%")
        print(f"  Taux erreur total: {100*(fp+fn)/total:.2f}%")

        return cm

    def generate_visualizations(self):
        """Générer les visualisations"""
        print("\n" + "="*80)
        print("📊 GÉNÉRATION DES VISUALISATIONS")
        print("="*80)

        fig = plt.figure(figsize=(18, 12))
        gs = fig.add_gridspec(3, 2, hspace=0.3, wspace=0.3)

        # 1. Matrice de confusion globale
        ax1 = fig.add_subplot(gs[0, 0])
        cm = confusion_matrix(self.y_true, self.y_pred)
        sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', ax=ax1,
                    xticklabels=['Non Fondée', 'Fondée'],
                    yticklabels=['Non Fondée', 'Fondée'],
                    cbar_kws={'label': 'Nombre de cas'})
        ax1.set_xlabel('Prédiction', fontweight='bold')
        ax1.set_ylabel('Réalité', fontweight='bold')
        ax1.set_title('Matrice de Confusion Globale', fontweight='bold', fontsize=12)

        # 2. Accuracy par famille
        ax2 = fig.add_subplot(gs[0, 1])
        top_families = self.results_by_family.head(10)
        bars = ax2.barh(range(len(top_families)), top_families['Accuracy_Globale'] * 100)
        ax2.set_yticks(range(len(top_families)))
        ax2.set_yticklabels(top_families['Famille'], fontsize=8)
        ax2.set_xlabel('Accuracy (%)', fontweight='bold')
        ax2.set_title('Accuracy par Famille (Top 10)', fontweight='bold', fontsize=12)
        ax2.grid(True, alpha=0.3, axis='x')

        # Colorer en fonction de l'accuracy
        for i, bar in enumerate(bars):
            acc = top_families.iloc[i]['Accuracy_Globale']
            if acc >= 0.98:
                bar.set_color('#2ecc71')
            elif acc >= 0.95:
                bar.set_color('#f39c12')
            else:
                bar.set_color('#e74c3c')

        # Ajouter les valeurs
        for i, val in enumerate(top_families['Accuracy_Globale'] * 100):
            ax2.text(val + 0.5, i, f'{val:.1f}%', va='center', fontsize=8)

        # 3. Distribution des erreurs par famille
        ax3 = fig.add_subplot(gs[1, 0])
        x = range(len(top_families))
        width = 0.35
        ax3.bar([i - width/2 for i in x], top_families['FP'], width,
                label='Faux Positifs (FP)', color='#e74c3c', alpha=0.7)
        ax3.bar([i + width/2 for i in x], top_families['FN'], width,
                label='Faux Négatifs (FN)', color='#e67e22', alpha=0.7)
        ax3.set_xticks(x)
        ax3.set_xticklabels(top_families['Famille'], rotation=45, ha='right', fontsize=8)
        ax3.set_ylabel('Nombre d\'erreurs', fontweight='bold')
        ax3.set_title('Distribution des Erreurs par Famille', fontweight='bold', fontsize=12)
        ax3.legend()
        ax3.grid(True, alpha=0.3, axis='y')

        # 4. Volume par famille
        ax4 = fig.add_subplot(gs[1, 1])
        ax4.bar(range(len(top_families)), top_families['N_Total'], color='#3498db', alpha=0.7)
        ax4.set_xticks(range(len(top_families)))
        ax4.set_xticklabels(top_families['Famille'], rotation=45, ha='right', fontsize=8)
        ax4.set_ylabel('Nombre de réclamations', fontweight='bold')
        ax4.set_title('Volume par Famille', fontweight='bold', fontsize=12)
        ax4.grid(True, alpha=0.3, axis='y')

        # 5. Precision vs Recall par famille
        ax5 = fig.add_subplot(gs[2, 0])
        scatter = ax5.scatter(self.results_by_family['Precision'] * 100,
                             self.results_by_family['Recall'] * 100,
                             s=self.results_by_family['N_Total']/10,
                             c=self.results_by_family['F1_Score'],
                             cmap='RdYlGn', alpha=0.6, edgecolors='black', linewidth=0.5)
        ax5.set_xlabel('Precision (%)', fontweight='bold')
        ax5.set_ylabel('Recall (%)', fontweight='bold')
        ax5.set_title('Precision vs Recall par Famille', fontweight='bold', fontsize=12)
        ax5.grid(True, alpha=0.3)

        # Ligne diagonale
        ax5.plot([0, 100], [0, 100], 'r--', alpha=0.3, label='Precision = Recall')
        ax5.legend()

        # Colorbar
        cbar = plt.colorbar(scatter, ax=ax5)
        cbar.set_label('F1-Score', fontweight='bold')

        # 6. Taux de fondées par famille
        ax6 = fig.add_subplot(gs[2, 1])
        ax6.bar(range(len(top_families)), top_families['Pct_Fondées'], color='#9b59b6', alpha=0.7)
        ax6.set_xticks(range(len(top_families)))
        ax6.set_xticklabels(top_families['Famille'], rotation=45, ha='right', fontsize=8)
        ax6.set_ylabel('% Fondées', fontweight='bold')
        ax6.set_title('Taux de Réclamations Fondées par Famille', fontweight='bold', fontsize=12)
        ax6.grid(True, alpha=0.3, axis='y')
        ax6.axhline(y=50, color='red', linestyle='--', alpha=0.5, label='50%')
        ax6.legend()

        plt.suptitle('ANALYSE DÉTAILLÉE PAR FAMILLE DE PRODUIT', fontsize=16, fontweight='bold', y=0.995)

        # Sauvegarder
        output_dir = Path('outputs/production/figures')
        output_dir.mkdir(parents=True, exist_ok=True)

        output_path = output_dir / 'family_analysis.png'
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"   ✅ family_analysis.png")

        plt.close()

    def save_excel_report(self):
        """Sauvegarder le rapport en Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HEADER_FONT = Font(color="FFFFFF", bold=True)
        SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Analyse par Famille"

        # Titre
        ws['A1'] = "ANALYSE PAR FAMILLE DE PRODUIT"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
        ws.merge_cells('A1:L1')

        # Headers
        row = 3
        headers = ['Famille', 'N Total', 'N Fondées', 'N Non Fondées', '% Fondées',
                   'Accuracy Global', 'Acc. Fondées', 'Acc. Non Fondées',
                   'Precision', 'Recall', 'F1-Score', 'Erreurs (FP+FN)']

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')

        # Données
        for r_idx, (_, data_row) in enumerate(self.results_by_family.iterrows(), row + 1):
            ws.cell(row=r_idx, column=1, value=data_row['Famille']).border = BORDER
            ws.cell(row=r_idx, column=2, value=int(data_row['N_Total'])).border = BORDER
            ws.cell(row=r_idx, column=3, value=int(data_row['N_Fondées'])).border = BORDER
            ws.cell(row=r_idx, column=4, value=int(data_row['N_Non_Fondées'])).border = BORDER

            cell = ws.cell(row=r_idx, column=5, value=data_row['Pct_Fondées'] / 100)
            cell.number_format = '0.0%'
            cell.border = BORDER

            cell = ws.cell(row=r_idx, column=6, value=data_row['Accuracy_Globale'])
            cell.number_format = '0.00%'
            cell.border = BORDER
            # Colorer selon accuracy
            if data_row['Accuracy_Globale'] >= 0.98:
                cell.fill = SUCCESS_FILL
            elif data_row['Accuracy_Globale'] >= 0.95:
                cell.fill = WARNING_FILL
            else:
                cell.fill = ERROR_FILL

            cell = ws.cell(row=r_idx, column=7, value=data_row['Accuracy_Fondées'])
            cell.number_format = '0.00%'
            cell.border = BORDER

            cell = ws.cell(row=r_idx, column=8, value=data_row['Accuracy_Non_Fondées'])
            cell.number_format = '0.00%'
            cell.border = BORDER

            cell = ws.cell(row=r_idx, column=9, value=data_row['Precision'])
            cell.number_format = '0.00%'
            cell.border = BORDER

            cell = ws.cell(row=r_idx, column=10, value=data_row['Recall'])
            cell.number_format = '0.00%'
            cell.border = BORDER

            cell = ws.cell(row=r_idx, column=11, value=data_row['F1_Score'])
            cell.number_format = '0.00%'
            cell.border = BORDER

            errors = int(data_row['FP'] + data_row['FN'])
            ws.cell(row=r_idx, column=12, value=errors).border = BORDER

        # Ajuster largeurs
        ws.column_dimensions['A'].width = 40
        for col in 'BCDEFGHIJKL':
            ws.column_dimensions[col].width = 15

        # Matrice de confusion
        ws_cm = wb.create_sheet("Matrice de Confusion")
        ws_cm['A1'] = "MATRICE DE CONFUSION GLOBALE"
        ws_cm['A1'].font = Font(bold=True, size=14)

        cm = confusion_matrix(self.y_true, self.y_pred)

        ws_cm['B3'] = "Prédit: Non Fondée"
        ws_cm['C3'] = "Prédit: Fondée"
        ws_cm['A4'] = "Réel: Non Fondée"
        ws_cm['A5'] = "Réel: Fondée"

        ws_cm['B4'] = cm[0, 0]
        ws_cm['C4'] = cm[0, 1]
        ws_cm['B5'] = cm[1, 0]
        ws_cm['C5'] = cm[1, 1]

        ws_cm['B4'].fill = SUCCESS_FILL
        ws_cm['C4'].fill = ERROR_FILL
        ws_cm['B5'].fill = WARNING_FILL
        ws_cm['C5'].fill = SUCCESS_FILL

        # Sauvegarder
        output_dir = Path('outputs/production')
        output_dir.mkdir(parents=True, exist_ok=True)

        output_path = output_dir / 'family_analysis.xlsx'
        wb.save(output_path)
        print(f"   ✅ family_analysis.xlsx")

    def run(self, y_true, y_pred, y_prob=None):
        """Exécution complète de l'analyse"""
        if not self.load_results():
            print("❌ Impossible de charger les données")
            return

        self.load_predictions(y_true, y_pred, y_prob)
        self.analyze_by_family()
        self.display_summary()
        self.generate_confusion_matrix_global()
        self.generate_visualizations()
        self.save_excel_report()

        print("\n" + "="*80)
        print("✅ ANALYSE TERMINÉE")
        print("="*80)
        print(f"\n📂 Résultats: outputs/production/")
        print("   - family_analysis.png")
        print("   - family_analysis.xlsx")


# Fonction utilitaire pour utiliser avec model_comparison.py
def analyze_from_comparison_results(model_name='XGBoost'):
    """
    Utilise les résultats de model_comparison.py pour faire l'analyse
    """
    import pickle

    # Essayer de charger les résultats
    try:
        # Si model_comparison.py a sauvegardé les résultats
        results_path = Path('outputs/production/models/comparison_results.pkl')
        if results_path.exists():
            with open(results_path, 'rb') as f:
                results = pickle.load(f)

            if model_name in results:
                analyzer = FamilyAnalyzer()
                analyzer.run(
                    y_true=results[model_name]['y_true'],
                    y_pred=results[model_name]['y_pred'],
                    y_prob=results[model_name]['y_prob']
                )
        else:
            print("⚠️  Fichier de résultats non trouvé. Exécutez d'abord model_comparison.py")
    except Exception as e:
        print(f"❌ Erreur: {e}")
        print("\nPour utiliser ce script, vous devez d'abord:")
        print("1. Exécuter model_comparison.py")
        print("2. Ou fournir manuellement y_true et y_pred:")
        print("\n   analyzer = FamilyAnalyzer()")
        print("   analyzer.run(y_true, y_pred, y_prob)")


if __name__ == '__main__':
    print("="*80)
    print("ANALYSE PAR FAMILLE DE PRODUIT")
    print("="*80)
    print("\nCe script doit être utilisé après model_comparison.py")
    print("\nExemple d'utilisation:")
    print("  from analyze_by_family import FamilyAnalyzer")
    print("  analyzer = FamilyAnalyzer()")
    print("  analyzer.run(y_true, y_pred, y_prob)")
